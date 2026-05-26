"""
=============================================================================
VRP com Janelas Temporais - Mercado Ibérico
=============================================================================
Todos os dados são lidos directamente do ficheiro Excel (folha "3.").
Basta corrigir o Excel — o código adapta-se automaticamente.

Estrutura da folha "3.":
  Linhas 7–18   → Matriz de distâncias (km)
  Linhas 27–38  → Matriz de tempos (min)
  Linhas 45–46  → Procuras diárias A e B (kg)
  Linhas 62–63  → Tempos de serviço A e B (min)
  Linhas 75–76  → Janelas temporais (min desde meia-noite)
  Linhas 83–84  → Dados dos veículos
  Linhas 91–102 → Reabastecimentos 55m³
  Linhas 109–120→ Reabastecimentos Semi-Reboque

Solver: PuLP + CBC  →  pip install pulp

Pressupostos (enunciado Parte 3):
  - O tempo de carregamento/descarregamento é de 2 min/palete (já incluído
    nos tempos de serviço lidos do Excel).
  - O tempo médio de abastecimento de um veículo é de 15 min.
  - É possível chegar mais cedo e esperar pelo início da janela de entrega.
  - A preparação do produto A termina às 02:00 (120 min).
  - A preparação do produto B termina às 04:00 (240 min).
  - Entregas de produto A e B não são feitas em simultâneo (modelos separados).
  - Split delivery permitido: a procura de um cliente pode ser dividida entre
    vários veículos (várias viagens). Necessário porque clientes como Barcelona
    (31.992 kg Produto A) excedem a capacidade de qualquer veículo disponível.
  - A dimensão da frota não está limitada pelo enunciado; o solver usa o
    mínimo necessário. O limite superior é calculado automaticamente a partir
    da procura total e da capacidade do veículo (com margem de segurança de +5).
  - Os clientes Évora, Santa Maria da Feira e Oviedo não recebem Produto B:
    a sua janela de entrega encerra antes de qualquer veículo conseguir chegar
    após o fim da preparação (04:00). Assume-se que estas localidades não
    contemplam o Produto B nas suas encomendas. Numa situação real, seria
    necessário renegociar a janela temporal com esses clientes.
=============================================================================
"""

import math
import pulp
from openpyxl import load_workbook
from pathlib import Path

# =============================================================================
# 0. CAMINHO PARA O EXCEL — ajuste aqui se necessário
# =============================================================================
EXCEL_PATH = Path(__file__).parent / "Tabelas_Parte_3.xlsx"
SHEET_NAME = "3."

# Constantes de layout da folha (linhas no Excel, 1-indexadas)
_ROW_DIST_START      = 7    # primeira linha de dados da matriz de distâncias
_ROW_TIME_START      = 27   # primeira linha de dados da matriz de tempos
_ROW_DEMAND_A        = 45   # linha com procuras do produto A
_ROW_DEMAND_B        = 46   # linha com procuras do produto B
_ROW_SERVICE_A       = 62   # linha com tempos de serviço A
_ROW_SERVICE_B       = 63   # linha com tempos de serviço B
_ROW_TW_EARLY        = 75   # linha com hora de início das janelas
_ROW_TW_LATE         = 76   # linha com hora de fim das janelas
_ROW_VEHICLE_55      = 83   # linha com dados do veículo 55m³
_ROW_VEHICLE_SR      = 84   # linha com dados do Semi-Reboque
_ROW_REFUEL55_START  = 91   # primeira linha da tabela de reabastecimentos 55m³
_ROW_REFUELSR_START  = 109  # primeira linha da tabela de reabastecimentos Semi-Reboque

_N_NODES             = 12   # 1 depósito + 11 clientes
_COL_DATA_START      = 5    # coluna E (1-indexada) → onde começam os valores numéricos
_COL_VEHICLE_COST_FIX = 23  # coluna W → Custos Fixos Diários (€)
_COL_VEHICLE_COST_VAR = 24  # coluna X → Custos Variáveis Diários (€/km)
_COL_VEHICLE_CAP      = 6   # coluna F → Capacidade (kg)

_DAYS_YEAR   = 264   # dias de trabalho/ano (célula C53 do Excel)
_REFUEL_TIME = 15    # minutos por abastecimento (enunciado)
_IP_A        = 120   # 02:00 em minutos — fim da preparação do produto A
_IP_B        = 240   # 04:00 em minutos — fim da preparação do produto B

# Margem de veículos extra acima do mínimo teórico necessário.
# O enunciado não impõe um limite de frota; este valor garante que o solver
# tem veículos suficientes para qualquer cenário.
_FLEET_MARGIN = 5

# Clientes que NÃO recebem Produto B por impossibilidade de cumprir a janela
# de entrega após o fim da preparação (04:00).
# Évora (nó 1), Santa Maria da Feira (nó 4), Oviedo (nó 7) — índices internos.
# Pressuposto: estas localidades não contemplam o Produto B nas suas encomendas.
_NODES_NO_PRODUCT_B = {1, 4, 7}


# =============================================================================
# 1. LEITURA DO EXCEL
# =============================================================================

def _read_matrix(ws, first_row: int, n: int) -> list:
    """
    Lê uma matriz n×n a partir de 'first_row' (linha Excel, 1-indexada).
    Os valores numéricos começam na coluna _COL_DATA_START.
    """
    matrix = []
    rows = list(ws.iter_rows(
        min_row=first_row, max_row=first_row + n - 1,
        min_col=_COL_DATA_START, max_col=_COL_DATA_START + n - 1,
        values_only=True,
    ))
    for row in rows:
        matrix.append([float(v) if v is not None else 0.0 for v in row])
    return matrix


def _read_row_values(ws, row: int, n_clients: int) -> list:
    """Lê n_clients valores numéricos a partir da coluna _COL_DATA_START."""
    cells = list(ws.iter_rows(
        min_row=row, max_row=row,
        min_col=_COL_DATA_START,
        max_col=_COL_DATA_START + n_clients - 1,
        values_only=True,
    ))[0]
    return [float(v) if v is not None else 0.0 for v in cells]


def _read_vehicle_row(ws, row: int) -> dict:
    """Lê uma linha de veículo e devolve um dicionário com os campos relevantes."""
    cells = list(ws.iter_rows(
        min_row=row, max_row=row,
        min_col=1, max_col=_COL_VEHICLE_COST_VAR,
        values_only=True,
    ))[0]
    # cells é 0-indexado a partir da col 1:
    #   cells[3]  = id          (col D)
    #   cells[5]  = capacity kg (col F)
    #   cells[22] = custo fixo diário €  (col W)
    #   cells[23] = custo variável €/km  (col X)
    return {
        "id":          int(cells[3]),
        "capacity_kg": float(cells[5]),
        "cost_fix":    float(cells[22]),
        "cost_var":    float(cells[23]),
    }


def load_data(excel_path: Path) -> dict:
    """
    Lê todos os dados do Excel e devolve um dicionário estruturado.
    Levanta FileNotFoundError com mensagem clara se o ficheiro não existir.
    """
    if not excel_path.exists():
        raise FileNotFoundError(
            f"\n[ERRO] Ficheiro Excel não encontrado: {excel_path}\n"
            f"Coloque 'Tabelas_Parte_3.xlsx' na mesma pasta que este script."
        )

    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    n  = _N_NODES
    nc = n - 1  # número de clientes

    dist = _read_matrix(ws, _ROW_DIST_START, n)
    time = _read_matrix(ws, _ROW_TIME_START, n)

    demand_a_list = _read_row_values(ws, _ROW_DEMAND_A, nc)
    demand_b_list = _read_row_values(ws, _ROW_DEMAND_B, nc)
    demand_a = {i + 1: demand_a_list[i] for i in range(nc)}
    # Évora (1), St. Mª Feira (4) e Oviedo (7) excluídos do Produto B:
    # a sua janela de entrega fecha antes de qualquer veículo poder chegar
    # após o fim da preparação (04:00). Pressuposto documentado no relatório.
    demand_b = {
        i + 1: (0.0 if (i + 1) in _NODES_NO_PRODUCT_B else demand_b_list[i])
        for i in range(nc)
    }

    svc_a_list = _read_row_values(ws, _ROW_SERVICE_A, nc)
    svc_b_list = _read_row_values(ws, _ROW_SERVICE_B, nc)
    service_a = {0: 0.0, **{i + 1: svc_a_list[i] for i in range(nc)}}
    service_b = {0: 0.0, **{i + 1: svc_b_list[i] for i in range(nc)}}

    tw_early_list = _read_row_values(ws, _ROW_TW_EARLY, nc)
    tw_late_list  = _read_row_values(ws, _ROW_TW_LATE,  nc)
    tw_early = {0: 0.0,    **{i + 1: tw_early_list[i] for i in range(nc)}}
    tw_late  = {0: 1440.0, **{i + 1: tw_late_list[i]  for i in range(nc)}}

    v55 = _read_vehicle_row(ws, _ROW_VEHICLE_55)
    vsr = _read_vehicle_row(ws, _ROW_VEHICLE_SR)
    vehicles = {
        v55["id"]: {
            "name":         "55m³",
            "capacity_kg":  v55["capacity_kg"],
            "cost_fix":     v55["cost_fix"],
            "cost_var":     v55["cost_var"],
            "refuel_table": _read_matrix(ws, _ROW_REFUEL55_START, n),
        },
        vsr["id"]: {
            "name":         "Semi-Reboque",
            "capacity_kg":  vsr["capacity_kg"],
            "cost_fix":     vsr["cost_fix"],
            "cost_var":     vsr["cost_var"],
            "refuel_table": _read_matrix(ws, _ROW_REFUELSR_START, n),
        },
    }

    wb.close()
    return {
        "dist":      dist,
        "time":      time,
        "demand_a":  demand_a,
        "demand_b":  demand_b,
        "service_a": service_a,
        "service_b": service_b,
        "tw_early":  tw_early,
        "tw_late":   tw_late,
        "vehicles":  vehicles,
    }


def compute_max_vehicles(data: dict, vehicle_ids: list, product: str) -> dict:
    """
    Calcula automaticamente o limite superior de veículos por tipo,
    com base na procura total e na capacidade de cada veículo.
    O enunciado não impõe um máximo de frota; este limite apenas garante
    que o solver tem veículos suficientes (mínimo teórico + _FLEET_MARGIN).
    """
    demand = data["demand_a"] if product == "A" else data["demand_b"]
    total_demand = sum(demand.values())
    max_vehicles = {}
    for vid in vehicle_ids:
        cap = data["vehicles"][vid]["capacity_kg"]
        min_needed = math.ceil(total_demand / cap)
        max_vehicles[vid] = min_needed + _FLEET_MARGIN
    return max_vehicles


# Nomes dos nós (ordem idêntica ao Excel)
NODE_NAMES = [
    "CD Saragoça", "Évora", "Faro", "Coimbra", "St Mª Feira",
    "Chaves", "Bragança", "Oviedo", "Toledo", "Valência",
    "Saragoça", "Barcelona",
]


# =============================================================================
# 2. AUXILIAR
# =============================================================================

def travel_time_total(i: int, j: int, vehicle_id: int, data: dict) -> float:
    """Tempo (min) de i→j incluindo paragens de abastecimento."""
    refuel = data["vehicles"][vehicle_id]["refuel_table"]
    return data["time"][i][j] + refuel[i][j] * _REFUEL_TIME


# =============================================================================
# 3. MODELO VRP-TW
# =============================================================================

def build_and_solve_vrp(
    product:     str,   # "A" ou "B"
    vehicle_ids: list,  # [1], [2] ou [1, 2]
    n_vehicles:  dict,  # {vehicle_id: n_max} — calculado automaticamente
    data:        dict,
    verbose:     bool = False,
) -> dict:
    """
    Modelo VRP com Janelas Temporais.

    Variáveis:
      x[i,j,k] ∈ {0,1}   veículo k percorre arco i→j
      b[i,k]   ≥ 0        instante de início de serviço no nó i pelo veículo k
      y[i,k]   ∈ [0,1]    fracção da procura de i entregue pelo veículo k (split delivery)
      z[k]     ∈ {0,1}    veículo k é utilizado

    Função objectivo:
      Min  Σ_k Σ_i Σ_j  dist[i][j] · x[i,j,k] · cv_k  +  Σ_k  z[k] · cf_k

    Restrições:
      (11) Σ_j x[0,j,k] ≤ 1                         ∀k
      (12) Σ_i x[i,p,k] = Σ_j x[p,j,k]              ∀p, ∀k
      (13) Σ_k y[i,k] = 1                            ∀i ∈ C
      (14) Σ_k Σ_i x[i,j,k] ≥ 1                     ∀j ∈ C
      (15) Σ_i q_i·y[i,k] ≤ cap_k                   ∀k
      (16) y[i,k] ≤ Σ_j x[i,j,k]                    ∀i ∈ C, ∀k
      (17) b[i,k]+s_i+t_ij−M(1−x[i,j,k]) ≤ b[j,k]  ∀i,j, ∀k
      (18) e_i ≤ b[i,k] ≤ l_i                        ∀i ∈ C, ∀k
      (19) b[0,k] ≥ IP                                ∀k
      (20) Σ_ij x[i,j,k] ≤ M'·z[k]                  ∀k
    """

    vehicles_data = data["vehicles"]
    demand  = data["demand_a"]  if product == "A" else data["demand_b"]
    service = data["service_a"] if product == "A" else data["service_b"]
    IP      = _IP_A             if product == "A" else _IP_B

    n         = _N_NODES
    all_nodes = list(range(n))
    clients   = list(range(1, n))
    active    = [i for i in clients if demand.get(i, 0) > 0]

    # Lista plana de veículos individuais
    veh_list = []
    for vid in vehicle_ids:
        for _ in range(n_vehicles[vid]):
            veh_list.append(vid)
    K = list(range(len(veh_list)))

    M_BIG   = 1440              # 24 h em minutos (big-M para restrição de sequência)
    M_PRIME = len(active) + 1   # big-M para restrição de activação de veículo

    prob = pulp.LpProblem(f"VRP_TW_{product}", pulp.LpMinimize)

    # --- Variáveis de decisão ---
    x = pulp.LpVariable.dicts(
        "x",
        [(i, j, k) for i in all_nodes for j in all_nodes for k in K if i != j],
        cat="Binary",
    )
    b = pulp.LpVariable.dicts(
        "b", [(i, k) for i in all_nodes for k in K],
        lowBound=0, cat="Continuous",
    )
    # y contínuo [0,1]: permite split delivery — a procura de um cliente pode ser
    # dividida entre vários veículos (várias viagens). Necessário porque alguns
    # clientes (ex: Barcelona, 31.992 kg) excedem a capacidade de qualquer veículo,
    # logo precisam obrigatoriamente de múltiplas visitas.
    y = pulp.LpVariable.dicts(
        "y", [(i, k) for i in active for k in K],
        lowBound=0, upBound=1, cat="Continuous",
    )
    z = pulp.LpVariable.dicts("z", K, cat="Binary")

    # --- Função objectivo ---
    vv = {k: vehicles_data[veh_list[k]]["cost_var"] for k in K}
    fv = {k: vehicles_data[veh_list[k]]["cost_fix"] for k in K}

    prob += (
        pulp.lpSum(
            data["dist"][i][j] * x[i, j, k] * vv[k]
            for i in all_nodes for j in all_nodes for k in K if i != j
        )
        + pulp.lpSum(fv[k] * z[k] for k in K),
        "Custo_Total",
    )

    # --- Restrições ---

    # (11) Cada veículo sai do depósito no máximo uma vez
    for k in K:
        prob += (
            pulp.lpSum(x[0, j, k] for j in active) <= 1,
            f"R11_k{k}",
        )

    # (12) Conservação de fluxo em cada nó
    for k in K:
        for p in all_nodes:
            prob += (
                pulp.lpSum(x[i, p, k] for i in all_nodes if i != p)
                == pulp.lpSum(x[p, j, k] for j in all_nodes if j != p),
                f"R12_p{p}_k{k}",
            )

    # (13) Toda a procura de cada cliente é satisfeita (fracções somam 1)
    for i in active:
        prob += (pulp.lpSum(y[i, k] for k in K) == 1, f"R13_i{i}")

    # (14) Cada cliente é visitado por pelo menos um veículo
    for j in active:
        prob += (
            pulp.lpSum(x[i, j, k] for i in all_nodes for k in K if i != j) >= 1,
            f"R14_j{j}",
        )

    # (15) Capacidade do veículo não é excedida
    for k in K:
        cap = vehicles_data[veh_list[k]]["capacity_kg"]
        prob += (
            pulp.lpSum(demand[i] * y[i, k] for i in active) <= cap,
            f"R15_k{k}",
        )

    # (16) Veículo só entrega num cliente se o visitar
    for k in K:
        for i in active:
            prob += (
                y[i, k] <= pulp.lpSum(x[i, j, k] for j in all_nodes if j != i),
                f"R16_i{i}_k{k}",
            )

    # (17) Propagação de tempo (garante sequência e janelas temporais)
    for k in K:
        for i in all_nodes:
            for j in active:
                if i != j:
                    t_ij = travel_time_total(i, j, veh_list[k], data)
                    s_i  = service.get(i, 0.0)
                    prob += (
                        b[i, k] + s_i + t_ij - M_BIG * (1 - x[i, j, k]) <= b[j, k],
                        f"R17_i{i}_j{j}_k{k}",
                    )

    # (18) Respeito pelas janelas temporais dos clientes
    for k in K:
        for i in active:
            prob += (b[i, k] >= data["tw_early"][i], f"R18a_i{i}_k{k}")
            prob += (b[i, k] <= data["tw_late"][i],  f"R18b_i{i}_k{k}")

    # (19) Veículos só saem após o fim da preparação do produto
    for k in K:
        prob += (b[0, k] >= IP, f"R19_k{k}")

    # (20) Activação de veículo: z[k]=1 se e só se o veículo k é usado
    for k in K:
        prob += (
            pulp.lpSum(x[i, j, k] for i in all_nodes for j in active if i != j)
            <= M_PRIME * z[k],
            f"R20_k{k}",
        )

    # --- Resolver ---
    solver = pulp.PULP_CBC_CMD(msg=1 if verbose else 0, timeLimit=300, gapRel=0.02)
    status = prob.solve(solver)

    # --- Extrair resultados ---
    result = {
        "status":          pulp.LpStatus[status],
        "obj":             pulp.value(prob.objective) or 0.0,
        "routes":          {},
        "n_vehicles_used": 0,
    }

    if pulp.LpStatus[status] not in ("Optimal", "Feasible"):
        return result

    for k in K:
        if pulp.value(z[k]) and pulp.value(z[k]) > 0.5:
            result["n_vehicles_used"] += 1

            # Clientes servidos por este veículo (y[i,k] > 0)
            client_loads = {
                i: round(demand[i] * (pulp.value(y[i, k]) or 0), 1)
                for i in active
                if (pulp.value(y[i, k]) or 0) > 0.01
            }

            # Construir rota a partir dos clientes com carga,
            # ordenados pelo instante de início de serviço b[i,k]
            served = sorted(
                client_loads.keys(),
                key=lambda i: pulp.value(b[i, k]) or 0
            )
            route = [0] + served + [0]
            load  = sum(client_loads.values())

            result["routes"][k] = {
                "vehicle_type": vehicles_data[veh_list[k]]["name"],
                "route_str":    " → ".join(NODE_NAMES[n] for n in route),
                "load_kg":      round(load, 1),
                "client_loads": client_loads,
            }

    return result


# =============================================================================
# 4. EXECUTAR CENÁRIOS (alíneas 3.1 a, b, c do enunciado)
# =============================================================================

def run_all_scenarios():
    print("A carregar dados do Excel...")
    data = load_data(EXCEL_PATH)
    vehicles = data["vehicles"]
    v_ids = sorted(vehicles.keys())   # [1, 2]

    scenarios = [
        (f"a) Apenas {vehicles[v_ids[0]]['name']:<14} | Produto A", [v_ids[0]], "A"),
        (f"a) Apenas {vehicles[v_ids[0]]['name']:<14} | Produto B", [v_ids[0]], "B"),
        (f"b) Apenas {vehicles[v_ids[1]]['name']:<14} | Produto A", [v_ids[1]], "A"),
        (f"b) Apenas {vehicles[v_ids[1]]['name']:<14} | Produto B", [v_ids[1]], "B"),
        (f"c) Frota Mista             | Produto A",                  v_ids,      "A"),
        (f"c) Frota Mista             | Produto B",                  v_ids,      "B"),
    ]

    print("=" * 74)
    print(f"  VRP-TW — MERCADO IBÉRICO  |  Dados: {EXCEL_PATH.name}")
    print("=" * 74)

    daily = {}

    for label, vids, prod in scenarios:
        print(f"\n{'─' * 74}")
        print(f"  {label}")
        print(f"{'─' * 74}")

        # Frota calculada automaticamente — sem limite artificial imposto
        n_veic = compute_max_vehicles(data, vids, prod)

        res = build_and_solve_vrp(prod, vids, n_veic, data, verbose=False)

        print(f"  Status        : {res['status']}")
        print(f"  Custo diário  : {res['obj']:>10.2f} €")
        print(f"  Custo anual   : {res['obj'] * _DAYS_YEAR:>10,.2f} €")
        print(f"  Nº veículos   : {res['n_vehicles_used']}")
        for _, r in res["routes"].items():
            print(f"    [{r['vehicle_type']:14s}]  {r['route_str']}  ({r['load_kg']:.0f} kg)")
            for node_idx, kg in r["client_loads"].items():
                pct = kg / r["load_kg"] * 100 if r["load_kg"] > 0 else 0
                print(f"        └─ {NODE_NAMES[node_idx]:<18} {kg:>8.1f} kg  ({pct:.0f}%)")

        key = (tuple(sorted(vids)), prod)
        daily[key] = daily.get(key, 0.0) + res["obj"]

    # --- Resumo anual consolidado ---
    print(f"\n{'=' * 74}")
    print("  RESUMO ANUAL — Produto A + B")
    print(f"{'=' * 74}")
    print(f"  {'Cenário':<32}  {'Diário (A+B)':>14}  {'Anual':>16}")
    print(f"  {'─' * 32}  {'─' * 14}  {'─' * 16}")

    for vids, lbl in [
        ([v_ids[0]], f"a) Apenas {vehicles[v_ids[0]]['name']}"),
        ([v_ids[1]], f"b) Apenas {vehicles[v_ids[1]]['name']}"),
        (v_ids,      "c) Frota Mista"),
    ]:
        ka  = (tuple(sorted(vids)), "A")
        kb  = (tuple(sorted(vids)), "B")
        dia = daily.get(ka, 0) + daily.get(kb, 0)
        ano = dia * _DAYS_YEAR
        print(f"  {lbl:<32}  {dia:>12.2f} €  {ano:>14,.2f} €")
    print()


if __name__ == "__main__":
    run_all_scenarios()